"""
Conversor de Excel para Markdown.
"""
from typing import Optional
import io
from loguru import logger

try:
    import pandas as pd
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def convert_excel_to_markdown(file_content: bytes, filename: str) -> Optional[str]:
    """
    Converte Excel para Markdown preservando estrutura de planilhas.
    
    Args:
        file_content: Conteúdo do arquivo em bytes
        filename: Nome do arquivo (para logging)
        
    Returns:
        str: Conteúdo Markdown ou None se erro
    """
    if not EXCEL_AVAILABLE:
        logger.error("Excel não suportado - pandas/openpyxl não instalados")
        return None
    
    try:
        excel_file = io.BytesIO(file_content)
        markdown_parts = [f"# {filename}\n\n"]
        
        # Carregar workbook para listar sheets
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        
        # Converter cada sheet
        for sheet_name in sheet_names:
            markdown_parts.append(f"## Planilha: {sheet_name}\n\n")
            
            # Ler sheet com pandas
            excel_file.seek(0)  # Resetar posição
            df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')
            
            if df.empty:
                markdown_parts.append("*Planilha vazia*\n\n")
                continue
            
            # Converter DataFrame para Markdown
            markdown_table = df.to_markdown(index=False, tablefmt="pipe")
            markdown_parts.append(markdown_table)
            markdown_parts.append("\n\n")
        
        workbook.close()
        return "".join(markdown_parts)
        
    except Exception as e:
        logger.error(f"Erro ao converter Excel '{filename}': {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
